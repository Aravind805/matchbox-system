from backend.app import app
from backend.models import db, Brand
from openpyxl import load_workbook
import os

BASE_DIR = os.path.dirname(__file__)
EXCEL_FILE = os.path.join(BASE_DIR, "Brand.xlsx")


#EXCEL_FILE = "Brand.xlsx"

with app.app_context():

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active   # first sheet

    # Expecting columns:
    # A = Brand
    # B = Sheets per kgs

    row_count = 0
    inserted = 0
    skipped = 0

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            # Skip header row
            continue

        brand_name = str(row[0]).strip()
        sheets_per_kg = row[1]

        if not brand_name or not sheets_per_kg:
            continue

        row_count += 1

        exists = Brand.query.filter_by(name=brand_name).first()
        if exists:
            skipped += 1
            print(f"ℹ️ Brand '{brand_name}' already exists")
            continue

        brand = Brand(
            name=brand_name,
            expected_sheets_per_kg=int(sheets_per_kg)
        )
        db.session.add(brand)
        inserted += 1
        print(f"✅ Brand '{brand_name}' added")

    db.session.commit()

    print("\n--- SUMMARY ---")
    print(f"Rows processed : {row_count}")
    print(f"Inserted       : {inserted}")
    print(f"Skipped        : {skipped}")
